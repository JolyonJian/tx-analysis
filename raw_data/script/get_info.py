from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup

wb = load_workbook(filename='.\\find.xlsx')
ws = wb['Sheet1']

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36', }

proxies = {"https": "http://127.0.0.1:10800", }


for i in range(1, 191):
    addr = ws['A' + str(i)].value
    url = 'https://etherscan.io/address/' + addr
    print('Get url %s.' % url)
    response = requests.get(url=url, headers=headers, proxies=proxies)
    soup = BeautifulSoup(response.text, 'html.parser')
    try:
        to = soup.select(
            '#transactions > div.table-responsive.mb-2.mb-md-0 > table > tbody > tr:nth-child(1) > td:nth-child(9) > span')
        tracer = soup.select(
            '#ContentPlaceHolder1_tr_tokeninfo > div > div.col-md-8 > a')
        to = (str(to).split('>')[1]).split('<')[0]
        tracer = (str(tracer).split('>')[1]).split('<')[0]
        ws['F' + str(i)] = to
        ws['G' + str(i)] = tracer
        print(addr, to, tracer)
    except:
        continue
wb.save('.\\find.xlsx')
