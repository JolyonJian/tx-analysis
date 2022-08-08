from openpyxl import load_workbook

addr_group = {}
addr_diff = {}

fr = open('.\\group_contract_list', 'r', encoding='utf8')
groups = fr.readlines()
count = 1
for group in groups:
    items = group.replace('\n', '').split(' ')
    master = (items[0].split('/')[-1]).split('_')[-1]
    m_addr = (master.split(':')[0]).split('.')[0]
    addr_group[m_addr] = count
    addr_diff[m_addr] = '-'
    if len(items) > 1:
        for item in items[1:]:
            info = (item.split('/')[-1]).split('_')[-1]
            addr = (info.split(':')[0]).split('.')[0]
            diff = (info.split(':')[1]).split('%')[0]
            if addr_group.get(addr) is None:
                addr_group[addr] = str(count)
                addr_diff[addr] = diff
            else:
                addr_group[addr] += ('/' + str(count))
                addr_diff[addr] += ('/' + diff)
    count += 1
# print(addr_group)
# print(addr_diff)

wb = load_workbook(filename='.\\contract_info.xlsx')
ws = wb['Contracts']
# Address: Bx Group: Gx Difference: Hx
for i in range(2, 9238):
    addr = ws['B' + str(i)].value
    if(addr_group.get(addr)):
        print(addr, addr_group[addr], addr_diff[addr])
        ws['G' + str(i)] = addr_group[addr]
        ws['H' + str(i)] = addr_diff[addr]
    else:
        print('Address %s is not in the groups.' % addr)
        continue
wb.save('.\\contract_info.xlsx')